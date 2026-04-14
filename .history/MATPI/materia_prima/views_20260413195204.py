from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import transaction, models
from django.utils import timezone
from .models import MateriaPrima, Lote
from usuarios.models import Administrador 
import openpyxl
from datetime import datetime
import re

# Función auxiliar para validar si el ID en sesión es Administrador
def check_admin(request):
    id_sesion = request.session.get('usuario_id')
    return Administrador.objects.filter(usuario_id=id_sesion).exists()

# --- VISTA PRINCIPAL (LISTADO) ---

def listar_materia_prima(request):
    id_sesion = request.session.get('usuario_id')
    if not id_sesion:
        return redirect('login')

    es_admin = check_admin(request)
    query = request.GET.get('buscar')
    
    if query:
        materia_primas = MateriaPrima.objects.filter(nombre_materia_prima__icontains=query)
    else:
        materia_primas = MateriaPrima.objects.all()
    
    return render(request, 'materia_prima/listar.html', {
        'materia_primas': materia_primas,
        'es_admin': es_admin,
        'buscar': query 
    })

# --- GESTIÓN DE MATERIA PRIMA (CREACIÓN ABIERTA A CAJERO Y ADMIN) ---

def mostrar_registro_materia_prima(request):
    id_sesion = request.session.get('usuario_id')
    if not id_sesion:
        return redirect('login')
    
    # Solo el administrador puede registrar materia prima
    es_admin = check_admin(request)
    if not es_admin:
        messages.error(request, "Solo el administrador puede registrar materia prima.")
        return redirect('listar_materia_prima')
        
    return render(request, 'materia_prima/registrar.html', {
        'es_admin': es_admin,
        'fecha_actual': timezone.now()
    })

def registrar_materia_prima(request):
    # Verificamos que sea administrador
    if not check_admin(request):
        messages.error(request, "No tienes permisos para realizar esta acción.")
        return redirect('listar_materia_prima')

    if request.method == 'POST':
        nombre = request.POST.get('txt_nombre')
        unidad = request.POST.get('txt_unidad')
        cantidad_unidad = int(request.POST.get('txt_cantidad_unidad', 1))
        tipo = request.POST.get('txt_tipo', 'Comida')
        cantidad = int(request.POST.get('txt_cantidad', 0))
        f_ingreso = request.POST.get('txt_fecha_ingreso')

        try:
            with transaction.atomic():
                materia = MateriaPrima.objects.create(
                    nombre_materia_prima=nombre,
                    unidad_medida=unidad,
                    cantidad_por_unidad=cantidad_unidad,
                    tipo=tipo,
                )
                
                if cantidad > 0:
                    Lote.objects.create(
                        materia_prima=materia,
                        cantidad_inicial=cantidad,
                        cantidad_actual=cantidad,
                        fecha_ingreso=f_ingreso or timezone.now(),
                        fecha_vencimiento=None
                    )
                
                messages.success(request, f"Materia prima '{nombre}' registrada exitosamente.")
        except Exception as e:
            messages.error(request, f"Error al registrar: {str(e)}")
            
        return redirect('listar_materia_prima')
    return redirect('mostrar_registro_materia_prima')

# --- EDICIÓN Y ELIMINACIÓN (SOLO ADMIN) ---

def pre_editar_materia_prima(request, id):
    es_admin = check_admin(request)
    # Aquí sí mantenemos el bloqueo de seguridad
    if not es_admin:
        messages.error(request, "Acceso denegado. Solo el administrador puede editar registros.")
        return redirect('listar_materia_prima')
        
    materia_prima = get_object_or_404(MateriaPrima, pk=id)
    return render(request, 'materia_prima/editar.html', {
        'materia_prima': materia_prima,
        'es_admin': es_admin
    })

def editar_materia_prima(request):
    if not check_admin(request):
        messages.error(request, "No tienes permisos para realizar esta acción.")
        return redirect('listar_materia_prima')

    if request.method == 'POST':
        id_materia = request.POST.get('txt_id')
        materia = get_object_or_404(MateriaPrima, pk=id_materia)
        
        materia.nombre_materia_prima = request.POST.get('txt_nombre')
        materia.unidad_medida        = request.POST.get('txt_unidad')
        materia.cantidad_por_unidad  = int(request.POST.get('txt_cantidad_unidad', 1))
        materia.tipo                 = request.POST.get('txt_tipo', 'Comida')
        materia.save()
        
        messages.success(request, "Información de la materia prima actualizada correctamente.")
        
    return redirect('listar_materia_prima')

def eliminar_materia_prima(request, id):
    if not check_admin(request):
        messages.error(request, "No tienes permisos para eliminar materia prima.")
        return redirect('listar_materia_prima')

    materia = get_object_or_404(MateriaPrima, pk=id)
    materia.delete()
    messages.success(request, "Materia prima eliminada correctamente.")
    return redirect('listar_materia_prima')


def ver_lotes(request, id_materia):
    id_sesion = request.session.get('usuario_id')
    if not id_sesion: return redirect('login')
    
    es_admin = check_admin(request)
    materia = get_object_or_404(MateriaPrima, pk=id_materia)
    # Lotes con stock disponible primero, luego por vencimiento
    lotes = materia.lotes.filter(cantidad_actual__gt=0).order_by('fecha_vencimiento')
    lotes_agotados = materia.lotes.filter(cantidad_actual=0).order_by('-fecha_ingreso')[:10]
    
    return render(request, 'materia_prima/lotes.html', {
        'materia': materia,
        'lotes': lotes,
        'lotes_agotados': lotes_agotados,
        'es_admin': es_admin,
        'fecha_actual': timezone.now()
    })

def pre_editar_lote(request, id_lote):
    if not check_admin(request):
        messages.error(request, "Acceso denegado. Solo administradores pueden editar lotes.")
        return redirect('listar_materia_prima')
        
    lote = get_object_or_404(Lote, pk=id_lote)
    return render(request, 'materia_prima/editar_lote.html', {
        'lote': lote,
        'es_admin': True
    })

def editar_lote(request):
    if not check_admin(request):
        messages.error(request, "Permiso denegado.")
        return redirect('listar_materia_prima')

    if request.method == 'POST':
        id_lote = request.POST.get('txt_id')
        lote = get_object_or_404(Lote, pk=id_lote)
        
        lote.cantidad_actual = int(request.POST.get('txt_cantidad', 0))
        lote.fecha_vencimiento = request.POST.get('txt_fecha_vencimiento') or None
        lote.save()
        
        messages.success(request, f"Lote #{lote.id} actualizado correctamente.")
        return redirect('ver_lotes', id_materia=lote.materia_prima.id)
    
    return redirect('listar_materia_prima')

def eliminar_lote(request, id_lote):
    if not check_admin(request):
        messages.error(request, "Permiso denegado.")
        return redirect('listar_materia_prima')

    lote = get_object_or_404(Lote, pk=id_lote)
    id_materia = lote.materia_prima.id
    lote.delete()
    
    messages.success(request, "Lote eliminado correctamente.")
    return redirect('ver_lotes', id_materia=id_materia)


# --- IMPORTACIÓN MASIVA DESDE EXCEL ---

def importar_materia_prima_excel(request):
    if not check_admin(request):
        messages.error(request, "No tienes permisos para realizar esta acción.")
        return redirect('listar_materia_prima')

    if request.method == 'POST' and request.FILES.get('archivo_excel'):
        archivo = request.FILES['archivo_excel']
        
        # Validar extensión
        if not archivo.name.endswith('.xlsx'):
            messages.error(request, "Solo se permiten archivos .xlsx")
            return redirect('importar_materia_prima_excel')

        try:
            wb = openpyxl.load_workbook(archivo)
            sheet = wb.active
            
            rows = list(sheet.iter_rows(min_row=2, values_only=True))
            
            # 1. Verificar si hay datos válidos y duplicados
            for row in rows:
                if len(row) < 4:
                    messages.error(request, "Los datos no son válidos")
                    return redirect('importar_materia_prima_excel')
                
                nombre, unidad, cant_unidad, tipo = row[:4]
                if not nombre: continue
                
                # Validar tipos de datos
                try:
                    cant_unidad = int(cant_unidad or 1)
                except (ValueError, TypeError):
                    messages.error(request, "Los datos no son válidos")
                    return redirect('importar_materia_prima_excel')
                
                unidad = unidad or 'und'
                tipo = tipo or 'Comida'

                if MateriaPrima.objects.filter(
                    nombre_materia_prima__iexact=nombre,
                    unidad_medida__iexact=unidad,
                    cantidad_por_unidad=cant_unidad,
                    tipo=tipo
                ).exists():
                    messages.error(request, "No se puede importar el archivo debido a que los datos ya existen o ya fueron ingresados")
                    return redirect('listar_materia_prima')

            # 2. Si no hay duplicados y todo es válido, importar
            creados = 0
            with transaction.atomic():
                for row in rows:
                    nombre, unidad, cant_unidad, tipo = row[:4]
                    if not nombre: continue
                    
                    MateriaPrima.objects.create(
                        nombre_materia_prima=nombre,
                        unidad_medida=unidad or 'und',
                        cantidad_por_unidad=int(cant_unidad or 1),
                        tipo=tipo or 'Comida'
                    )
                    creados += 1
            
            if creados > 0:
                messages.success(request, f"Se importaron {creados} materias primas correctamente.")
                
        except Exception as e:
            messages.error(request, f"Error al procesar el archivo: {str(e)}")
            
        return redirect('listar_materia_prima')
        
    return render(request, 'materia_prima/importar_materia_prima.html')


def importar_lotes_excel(request):
    if not check_admin(request):
        messages.error(request, "No tienes permisos para realizar esta acción.")
        return redirect('listar_materia_prima')

    if request.method == 'POST' and request.FILES.get('archivo_excel'):
        archivo = request.FILES['archivo_excel']

        # Validar extensión
        if not archivo.name.endswith('.xlsx'):
            messages.error(request, "Solo se permiten archivos .xlsx")
            return redirect('importar_lotes_excel')

        try:
            wb = openpyxl.load_workbook(archivo)
            sheet = wb.active
            
            rows = list(sheet.iter_rows(min_row=2, values_only=True))
            lotes_a_crear = []

            # 1. Validación de tipos, existencia de materias y duplicados
            for row_idx, row in enumerate(rows, start=2):
                if len(row) < 2:
                    messages.error(request, "Los datos no son válidos")
                    return redirect('importar_lotes_excel')

                nombre_completo, cantidad, f_vencimiento, precio = row[:4]
                if not nombre_completo or cantidad is None: continue
                
                # Validar tipos numéricos
                try:
                    cantidad = float(cantidad)
                    if precio is not None:
                        precio = float(precio)
                except (ValueError, TypeError):
                    messages.error(request, "Los datos no son válidos")
                    return redirect('importar_lotes_excel')

                nombre_completo = str(nombre_completo).strip()
                materia = None
                
                materia = MateriaPrima.objects.filter(nombre_materia_prima__iexact=nombre_completo).first()
                if not materia:
                    if match := re.match(r"^(.*?)\s*\(([\d.,]+)\s+(.*)\)$", nombre_completo):
                        nombre_base = match.group(1).strip()
                        equiv_str = match.group(2).replace(',', '.')
                        unidad_str = match.group(3).strip()
                        try:
                            equiv_int = int(float(equiv_str))
                            materia = MateriaPrima.objects.filter(
                                nombre_materia_prima__iexact=nombre_base,
                                cantidad_por_unidad=equiv_int,
                                unidad_medida__iexact=unidad_str
                            ).first()
                        except: pass
                
                if not materia:
                    messages.error(request, f"Error en fila {row_idx}: Materia prima '{nombre_completo}' no encontrada.")
                    return redirect('importar_lotes_excel')

                # Validar fecha
                if isinstance(f_vencimiento, str):
                    try:
                        f_vencimiento = datetime.strptime(f_vencimiento, '%Y-%m-%d').date()
                    except:
                        f_vencimiento = None
                elif isinstance(f_vencimiento, datetime):
                    f_vencimiento = f_vencimiento.date()

                # VERIFICAR DUPLICADO EN DB
                if Lote.objects.filter(
                    materia_prima=materia,
                    cantidad_inicial=cantidad,
                    fecha_vencimiento=f_vencimiento,
                    precio_unidad=precio
                ).exists():
                    messages.error(request, "No se puede importar el archivo debido a que los datos ya existen")
                    return redirect('listar_materia_prima')

                lotes_a_crear.append({
                    'materia_prima': materia,
                    'cantidad_inicial': cantidad,
                    'fecha_vencimiento': f_vencimiento,
                    'precio_unidad': precio
                })

            # 2. Si todo es válido, crear los lotes
            if lotes_a_crear:
                with transaction.atomic():
                    for data in lotes_a_crear:
                        Lote.objects.create(
                            materia_prima=data['materia_prima'],
                            cantidad_inicial=data['cantidad_inicial'],
                            cantidad_actual=data['cantidad_inicial'],
                            fecha_ingreso=timezone.now(),
                            fecha_vencimiento=data['fecha_vencimiento'],
                            precio_unidad=data['precio_unidad']
                        )
                messages.success(request, f"Se importaron {len(lotes_a_crear)} lotes correctamente.")
                
        except Exception as e:
            messages.error(request, f"Error al procesar el archivo: {str(e)}")
            
        return redirect('listar_materia_prima')
        
    return render(request, 'materia_prima/importar_lotes.html')

